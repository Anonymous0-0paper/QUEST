import random

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Font

from algorithms.Experiment import Experiment
from algorithms.Fuzzy import Fuzzy
from algorithms.Greedy import Greedy
from algorithms.MOPSO import MOPSO
from algorithms.MQGA import MQGA
from algorithms.NSGA3 import NSGA3
from algorithms.QUEST import QUEST
from algorithms.Random import Random
from network.Network import Network

loads = [50, 100, 150]
dag_files = {
    50: [f"./workflow/Outputs/montage-50/dag-{i + 1}.json" for i in range(100)],
    100: [f"./workflow/Outputs/montage-100/dag-{i + 1}.json" for i in range(100)],
    150: [f"./workflow/Outputs/montage-150/dag-{i + 1}.json" for i in range(100)],
}
output = f"./motivation-1.xlsx"

samples = 10
algorithms = [Random, QUEST, NSGA3, MQGA, MOPSO, Fuzzy, Greedy]

if __name__ == '__main__':
    results: list[dict[str, float]] = []

    for sample in range(samples):
        Alg = random.choice(algorithms)
        Experiment.progress(sample / samples, Alg.__name__)

        load = random.choice(loads)

        file = random.choice(dag_files[load])
        dag = Experiment.read_dag(file)
        network = Network.generate()

        alg = Alg(network, dag)
        alg.run()

        value = {
            "Data Age": alg.calculate_data_age(),
            "Latency": alg.calculate_latency(),
            "Energy": alg.calculate_energy(),
        }
        results.append(value)

        Experiment.progress((sample + 1) / samples, Alg.__name__)

    wb = load_workbook(output)

    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_font = Font(bold=True)

    ws = Experiment.create_sheet(wb, "data age")

    row = 1
    for header in ["Data Age", "Latency", "Energy"]:
        cell = ws.cell(row, 1, header)
        cell.fill = header_fill
        cell.font = header_font

        for sample in range(samples):
            ws.cell(row, sample + 2, results[sample][header])

        row += 1

    wb.save(output)
