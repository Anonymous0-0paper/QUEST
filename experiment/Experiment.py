import csv
import json
import os
import sys
from statistics import mean
from time import time
from psutil import cpu_percent, virtual_memory, disk_usage

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.workbook import Workbook

from algorithms.Algorithm import Algorithm
from algorithms.Fuzzy import Fuzzy
from algorithms.Greedy import Greedy
from algorithms.MOPSO import MOPSO
from algorithms.MQGA import MQGA
from algorithms.NSGA3 import NSGA3
from algorithms.QUEST import QUEST
from algorithms.RL import RL
from algorithms.Random import Random
from model.DAGModel import DAGModel
from model.SubTaskModel import SubTaskModel
from network.Network import Network
from network.Node import NodeType


class Experiment:

    def __init__(self, algorithms: list[str], loads: list[int], dag_files: dict[int, list[str]], iteration: int,
                 output: str,path):
        self.algorithms = algorithms
        self.loads = loads
        self.dag_files = dag_files
        self.iteration = iteration
        self.output = output
        self.path=path


        # key: (algorithm, load, metric_index)
        # value: list of samples
        self.result: dict[tuple[str, int, int], list[float]] = {}
        self.result_load: dict[tuple[str, int], list[list[float]]] = {}
        self.utilization_data = []  # To store CPU, memory, and disk usage
        self.execution_times = []  # To store execution times
        self.peak_cpu_data = []    # To store peak CPU usage
        self.wb = load_workbook(self.output)
        self.network = Network.generate(path=self.path)

    def run(self):
        total = len(self.loads) * len(self.algorithms) * self.iteration
        for l in range(len(self.loads)):
            load = self.loads[l]

            for i in range(self.iteration):
                file = self.dag_files[load][i % len(self.dag_files[load])]
                dag = Experiment.read_dag(file)

                for a in range(len(self.algorithms)):
                    algorithm = self.algorithms[a]

                    current: float = (
                            l * len(self.algorithms) * self.iteration +
                            i * len(self.algorithms) +
                            a
                    )
                    self.progress(current / total, algorithm)

                    if algorithm == "Random":
                        alg = Random(self.network, dag)
                    elif algorithm == "Fuzzy":
                        alg = Fuzzy(self.network, dag)
                    elif algorithm == "NSGA3":
                        alg = NSGA3(self.network, dag)
                    elif algorithm == "QUEST":
                        alg = QUEST(self.network, dag)
                    elif algorithm == "QUEST_NDVFS":
                        alg = QUEST(self.network, dag, False)
                    elif algorithm == "MQGA":
                        alg = MQGA(self.network, dag)
                    elif algorithm == "Greedy":
                        alg = Greedy(self.network, dag)
                    elif algorithm == "MOPSO":
                        alg = MOPSO(self.network, dag)
                    elif algorithm == "RL":
                        alg = RL(self.network, dag)
                    else:
                        raise ValueError(f"Unknown algorithm: {algorithm}")


                    # Record resource utilization and execution time
                    start_time = time()
                    cpu_before = cpu_percent(interval=None)
                    memory_before = virtual_memory().percent
                    disk_before = disk_usage('/').percent

                    alg.run()

                    cpu_after = cpu_percent(interval=None)
                    memory_after = virtual_memory().percent
                    disk_after = disk_usage('/').percent
                    end_time = time()

                    execution_time = end_time - start_time
                    self.execution_times.append((algorithm, load, execution_time))
                    self.utilization_data.append((
                        algorithm, load,
                        (cpu_before + cpu_after) / 2,
                        (memory_before + memory_after) / 2,
                        (disk_before + disk_after) / 2
                    ))

                    if i == 0:
                        # Initialize lists for each metric
                        for metric_idx in range(6):
                            self.result[(algorithm, load, metric_idx)] = []
                        self.result_load[(algorithm, load)] = []

                    # Append results for each metric
                    self.result[(algorithm, load, 0)].append(alg.calculate_data_age())
                    self.result[(algorithm, load, 1)].append(alg.calculate_energy())
                    self.result[(algorithm, load, 2)].append(alg.calculate_completion_time())
                    self.result[(algorithm, load, 3)].append(alg.calculate_success())
                    (load_edge, load_cloud) = alg.calculate_load()
                    self.result[(algorithm, load, 4)].append(load_edge)
                    self.result[(algorithm, load, 5)].append(load_cloud)
                    self.result_load[(algorithm, load)].append(alg.calculate_load_per_node())

                    self.progress((current + 1) / total, algorithm)

    def store(self):
        metrics = {
            "Data Age": 0,
            "Energy": 1,
            "Makespan": 2,
            "Success Rate": 3,
            "Load Edge": 4,
            "Load Cloud": 5,
        }

        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        header_font = Font(bold=True)

        for metric_name, metric_index in metrics.items():
            ws = self.create_sheet(self.wb, metric_name)

            # Write headers
            headers = ["Algorithm", "Load", "Average", "Min", "Max", "Std Dev"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font

            row = 2
            for algorithm in self.algorithms:
                for load in self.loads:
                    values = self.result.get((algorithm, load, metric_index), [])
                    if not values:
                        continue  # Skip if no data

                    avg = mean(values)
                    min_val = min(values)
                    max_val = max(values)
                    variance = sum((x - avg) ** 2 for x in values) / len(values)
                    std_dev = variance ** 0.5

                    ws.cell(row=row, column=1, value=algorithm)
                    ws.cell(row=row, column=2, value=load)
                    ws.cell(row=row, column=3, value=avg)
                    ws.cell(row=row, column=4, value=min_val)
                    ws.cell(row=row, column=5, value=max_val)
                    ws.cell(row=row, column=6, value=std_dev)

                    row += 1

            # Add summary section
            row += 2
            summary_header = ws.cell(row=row, column=1, value="Summary by Algorithm")
            summary_header.font = header_font
            row += 1

            for algorithm in self.algorithms:
                all_values = []
                for load in self.loads:
                    all_values.extend(self.result.get((algorithm, load, metric_index), []))

                if not all_values:
                    continue  # Skip if no data

                avg = mean(all_values)
                min_val = min(all_values)
                max_val = max(all_values)
                variance = sum((x - avg) ** 2 for x in all_values) / len(all_values)
                std_dev = variance ** 0.5

                ws.cell(row=row, column=1, value=algorithm)
                ws.cell(row=row, column=3, value=avg)
                ws.cell(row=row, column=4, value=min_val)
                ws.cell(row=row, column=5, value=max_val)
                ws.cell(row=row, column=6, value=std_dev)

                row += 1

            row += 2
            chart_header_1 = ws.cell(row=row, column=1, value="Average data for chart")
            chart_header_1.font = header_font
            row += 1

            ws.cell(row=row, column=1, value="Algorithms")
            for l in range(len(self.loads)):
                ws.cell(row=row, column=l + 2, value=self.loads[l])

            row += 1

            for algorithm in self.algorithms:
                ws.cell(row=row, column=1, value=algorithm)

                for l in range(len(self.loads)):
                    avg = mean(self.result.get((algorithm, self.loads[l], metric_index)))
                    ws.cell(row=row, column=l + 2, value=avg)

                row += 1

            row += 2
            chart_header_2 = ws.cell(row=row, column=1, value="All data for chart")
            chart_header_2.font = header_font
            row += 1

            ws.cell(row=row, column=1, value="Algorithms")
            for l in range(len(self.loads)):
                for i in range(self.iteration):
                    ws.cell(row=row, column=l * self.iteration + i + 2, value=self.loads[l])

            row += 1

            for algorithm in self.algorithms:
                ws.cell(row=row, column=1, value=algorithm)

                for l in range(len(self.loads)):
                    for i in range(self.iteration):
                        data = self.result.get((algorithm, self.loads[l], metric_index))[i]
                        ws.cell(row=row, column=l * self.iteration + i + 2, value=data)

                row += 1

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Save node level load in edges
        ws = self.create_sheet(self.wb, "Loads")
        row = 1
        chart_header_1 = ws.cell(row=row, column=1, value="Average data for chart")
        chart_header_1.font = header_font
        row += 1

        for l in range(len(self.loads)):
            for n in range(len(self.network.nodes)):
                col = l * len(self.network.nodes) + n + 2
                node_type = "Edge" if self.network.nodes[n].node_type == NodeType.Edge else "Cloud"
                print(row, col, self.loads[l])
                ws.cell(row=row, column=col, value=node_type)
                ws.cell(row=row + 1, column=col, value=self.loads[l])
                ws.cell(row=row + 2, column=col, value=n + 1)

        row += 2
        ws.cell(row=row, column=1, value="Algorithms")
        row += 1

        for algorithm in self.algorithms:
            ws.cell(row=row, column=1, value=algorithm)

            for l in range(len(self.loads)):
                for n in range(len(self.network.nodes)):
                    col = l * len(self.network.nodes) + n + 2
                    total = 0
                    for i in range(self.iteration):
                        total += self.result_load.get((algorithm, self.loads[l]))[i][n]
                    ws.cell(row=row, column=col, value=total / len(self.network.nodes))

            row += 1


        self.wb.save(self.output)

        # --- Added Code for Saving Data to CSVs ---
        self.store_csvs(metrics)
        self.store_utilization_csv()
        self.store_execution_time_csv()

        # --- Compare QUEST with others and print ---
        self.compare_quest(metrics)

        # --- Store improvements in CSV ---
        self.store_improvements_csv(metrics)

    def store_utilization_csv(self):
        """
        Saves CPU, memory, and disk utilization data into a CSV file.
        """
        csv_path = os.path.join(os.path.dirname(self.output), "utilization_data.csv")
        with open(csv_path, mode='w', newline='') as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(["Algorithm", "Load", "CPU (%)", "Memory (%)", "Disk (%)"])
            writer.writerows(self.utilization_data)
        print(f"Utilization data saved to {csv_path}")

    def store_execution_time_csv(self):
        """
        Saves algorithm execution times into a CSV file.
        """
        csv_path = os.path.join(os.path.dirname(self.output), "execution_times.csv")
        with open(csv_path, mode='w', newline='') as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(["Algorithm", "Load", "Execution Time (s)"])
            writer.writerows(self.execution_times)
        print(f"Execution times saved to {csv_path}")

    def store_csvs(self, metrics: dict):
        """
        Stores the average, min, and max for each metric into separate CSV files.
        Each CSV file corresponds to a single metric and contains the Algorithm, Load, Average, Min, and Max.
        """
        # Define directory to store CSVs
        csv_dir = os.path.join(os.path.dirname(self.output), "csv_results")
        os.makedirs(csv_dir, exist_ok=True)

        for metric_name, metric_index in metrics.items():
            csv_filename = f"{metric_name.replace(' ', '_')}.csv"
            csv_path = os.path.join(csv_dir, csv_filename)

            with open(csv_path, mode='w', newline='') as csv_file:
                writer = csv.writer(csv_file)
                # Write headers
                writer.writerow(["Algorithm", "Load", "Average", "Min", "Max"])

                for algorithm in self.algorithms:
                    for load in self.loads:
                        values = self.result.get((algorithm, load, metric_index), [])
                        if not values:
                            continue  # Skip if no data

                        avg = mean(values)
                        min_val = min(values)
                        max_val = max(values)

                        writer.writerow([algorithm, load, avg, min_val, max_val])

        print(f"CSV files have been saved in the directory: {csv_dir}")

    def compare_quest(self, metrics: dict):
        """
        Compares the QUEST algorithm with other algorithms and prints the differences.
        """
        print("\n--- QUEST vs Other Algorithms Comparison ---\n")
        for metric_name, metric_index in metrics.items():
            print(f"Metric: {metric_name}")
            try:
                # Aggregate QUEST metrics across all loads
                quest_values = []
                for load in self.loads:
                    quest_values.extend(self.result[("QUEST", load, metric_index)])
                quest_avg = mean(quest_values)
                quest_min = min(quest_values)
                quest_max = max(quest_values)

                print(f"  QUEST - Average: {quest_avg:.4f}, Min: {quest_min:.4f}, Max: {quest_max:.4f}")

                # Determine if lower is better or higher is better
                lower_better_metrics = {"Data Age", "Energy", "Makespan", "Load"}
                higher_better_metrics = {"Success Rate"}

                for algorithm in self.algorithms:
                    if algorithm == "QUEST":
                        continue  # Skip QUEST itself

                    other_values = []
                    for load in self.loads:
                        other_values.extend(self.result.get((algorithm, load, metric_index), []))
                    if not other_values:
                        print(f"  {algorithm} has no data for this metric.")
                        continue
                    other_avg = mean(other_values)
                    other_min = min(other_values)
                    other_max = max(other_values)

                    if metric_name in lower_better_metrics:
                        avg_diff = other_avg - quest_avg
                        min_diff = other_min - quest_min
                        max_diff = other_max - quest_max
                        better_avg = "QUEST is better" if avg_diff > 0 else "Other is better"
                        better_min = "QUEST is better" if min_diff > 0 else "Other is better"
                        better_max = "QUEST is better" if max_diff > 0 else "Other is better"

                        # Percentage improvement (relative to QUEST)
                        # For lower-better: improvement% = ((other - quest)/quest)*100
                        if quest_avg != 0:
                            avg_improvement_pct = ((other_avg - quest_avg) / quest_avg) * 100
                        else:
                            avg_improvement_pct = float('inf')

                    elif metric_name in higher_better_metrics:
                        avg_diff = quest_avg - other_avg
                        min_diff = quest_min - other_min
                        max_diff = quest_max - other_max
                        better_avg = "QUEST is better" if avg_diff > 0 else "Other is better"
                        better_min = "QUEST is better" if min_diff > 0 else "Other is better"
                        better_max = "QUEST is better" if max_diff > 0 else "Other is better"

                        # For higher-better: improvement% = ((quest - other)/other)*100
                        if other_avg != 0:
                            avg_improvement_pct = ((quest_avg - other_avg) / other_avg) * 100
                        else:
                            avg_improvement_pct = float('inf')
                    else:
                        # Default to lower-better logic if not specified
                        avg_diff = other_avg - quest_avg
                        if quest_avg != 0:
                            avg_improvement_pct = ((other_avg - quest_avg) / quest_avg) * 100
                        else:
                            avg_improvement_pct = float('inf')

                    print(
                        f"  {algorithm} - Average Difference: {avg_diff:.4f} ({better_avg}), Improvement: {avg_improvement_pct:.2f}%")
                    print(f"             Min Difference: {min_diff:.4f} ({better_min})")
                    print(f"             Max Difference: {max_diff:.4f} ({better_max})")
            except KeyError:
                print("  QUEST data not available for this metric.")
            print()

    def store_improvements_csv(self, metrics: dict):
        """
        Creates CSV files that show improvements in percentage compared to QUEST for each algorithm.
        For lower-better metrics: Improvement% = ((OtherAvg - QuestAvg) / QuestAvg) * 100
        For higher-better metrics: Improvement% = ((QuestAvg - OtherAvg) / OtherAvg) * 100
        """
        csv_dir = os.path.join(os.path.dirname(self.output), "csv_results")
        os.makedirs(csv_dir, exist_ok=True)

        lower_better_metrics = {"Data Age", "Energy", "Makespan", "Load"}
        higher_better_metrics = {"Success Rate"}

        for metric_name, metric_index in metrics.items():
            # Check if QUEST data is available
            quest_values = []
            for load in self.loads:
                if ("QUEST", load, metric_index) in self.result:
                    quest_values.extend(self.result[("QUEST", load, metric_index)])
            if not quest_values:
                # No QUEST data, skip
                continue

            quest_avg = mean(quest_values)

            csv_filename = f"{metric_name.replace(' ', '_')}_improvements.csv"
            csv_path = os.path.join(csv_dir, csv_filename)

            with open(csv_path, mode='w', newline='') as csv_file:
                writer = csv.writer(csv_file)
                # Write headers
                writer.writerow(["Algorithm", "Load", "QuestAvg", "OtherAvg", "AvgDiff", "Improvement(%)"])

                for algorithm in self.algorithms:
                    if algorithm == "QUEST":
                        continue

                    for load in self.loads:
                        values = self.result.get((algorithm, load, metric_index), [])
                        if not values:
                            continue
                        other_avg = mean(values)

                        if metric_name in lower_better_metrics:
                            # Improvement% = ((other_avg - quest_avg)/quest_avg)*100
                            if quest_avg != 0:
                                improvement_pct = ((other_avg - quest_avg) / quest_avg) * 100
                            else:
                                improvement_pct = float('inf')
                            avg_diff = other_avg - quest_avg

                        elif metric_name in higher_better_metrics:
                            # Improvement% = ((quest_avg - other_avg)/other_avg)*100
                            if other_avg != 0:
                                improvement_pct = ((quest_avg - other_avg) / other_avg) * 100
                            else:
                                improvement_pct = float('inf')
                            avg_diff = quest_avg - other_avg

                        else:
                            # Default to lower-better
                            if quest_avg != 0:
                                improvement_pct = ((other_avg - quest_avg) / quest_avg) * 100
                            else:
                                improvement_pct = float('inf')
                            avg_diff = other_avg - quest_avg

                        writer.writerow([algorithm, load, quest_avg, other_avg, avg_diff, improvement_pct])

    @staticmethod
    def create_sheet(wb: Workbook, sheet_name: str):
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
        return ws

    @staticmethod
    def progress(percent: float, algorithm: str):
        arrow = '=' * int(round(percent * 100) - 1)
        spaces = ' ' * (100 - len(arrow))
        sys.stdout.write(f'\rProgress: [{arrow + spaces}] {int(percent * 100)}% [{algorithm}]')
        sys.stdout.flush()

    @staticmethod
    def read_dag(file: str):
        with open(file, 'r') as file:
            json_data = json.load(file)
            subtasks: list[SubTaskModel] = []
            for subtask_dto in json_data["subtasks"]:
                subtask = SubTaskModel(
                    subtask_dto["id"],
                    subtask_dto["executionCost"],
                    subtask_dto["memory"]
                )
                subtasks.append(subtask)

            dag = DAGModel(subtasks, json_data["edges"], json_data["deadline"])
            return dag
