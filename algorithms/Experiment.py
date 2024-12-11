import json
import sys
import csv
import os
from statistics import mean

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Font

from algorithms.Algorithm import Algorithm
from algorithms.Fuzzy import Fuzzy
from algorithms.NSGA3 import NSGA3
from algorithms.QUEST import QUEST
from algorithms.Random import Random
from model.DAGModel import DAGModel
from model.SubTaskModel import SubTaskModel
from network.Network import Network


class Experiment:

    def __init__(self, algorithms: list[str], loads: list[int], dag_files: dict[int, list[str]], iteration: int,
                 output: str):
        self.algorithms = algorithms
        self.loads = loads
        self.dag_files = dag_files
        self.iteration = iteration
        self.output = output

        # key: (algorithm, load, metric_index)
        # value: list of samples
        self.result: dict[tuple[str, int, int], list[float]] = {}
        self.wb = load_workbook(self.output)

    def run(self):
        network = Network.generate()

        total = len(self.loads) * len(self.algorithms) * self.iteration
        for l in range(len(self.loads)):
            load = self.loads[l]

            for i in range(self.iteration):
                file_index = i % len(self.dag_files[load])
                dags: DAGModel | None = None
                with open(self.dag_files[load][file_index], 'r') as file:
                    json_data = json.load(file)
                    subtasks: list[SubTaskModel] = []
                    for subtask_dto in json_data["subtasks"]:
                        subtask = SubTaskModel(
                            subtask_dto["id"],
                            subtask_dto["executionCost"],
                            subtask_dto["memory"]
                        )
                        subtasks.append(subtask)

                    dag = DAGModel(subtasks, json_data["edges"], json_data["deadline"])

                for a in range(len(self.algorithms)):
                    algorithm = self.algorithms[a]

                    current: float = (
                            l * len(self.algorithms) * self.iteration +
                            i * len(self.algorithms) +
                            a
                    )

                    alg: Algorithm | None = None
                    if algorithm == "Random":
                        alg = Random(network, dag)
                    elif algorithm == "Fuzzy":
                        alg = Fuzzy(network, dag)
                    elif algorithm == "NSGA3":
                        alg = NSGA3(network, dag)
                    elif algorithm == "QUEST":
                        alg = QUEST(network, dag)

                    alg.run()

                    if i == 0:
                        # Initialize lists for each metric
                        for metric_idx in range(5):
                            self.result[(algorithm, load, metric_idx)] = []

                    # Append results for each metric
                    self.result[(algorithm, load, 0)].append(alg.calculate_data_age())
                    self.result[(algorithm, load, 1)].append(alg.calculate_energy())
                    self.result[(algorithm, load, 2)].append(alg.calculate_completion_time())
                    self.result[(algorithm, load, 3)].append(alg.calculate_load_balance())
                    self.result[(algorithm, load, 4)].append(alg.calculate_success())

                    self.progress((current + 1) / total, algorithm)

    def store(self):
        metrics = {
            "Data Age": 0,
            "Energy": 1,
            "Makespan": 2,
            "Load": 3,
            "Success Rate": 4
        }

        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        header_font = Font(bold=True)

        for metric_name, metric_index in metrics.items():
            ws = self.create_sheet(metric_name)

            # Write headers
            headers = ["Algorithm", "Load", "Average", "Min", "Max", "Std Dev"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font

            row = 2
            for algorithm in self.algorithms:
                for load in self.loads:
                    values = self.result.get((algorithm, load, metric_index), [])
                    if not values:
                        continue  # Skip if no data

                    avg = mean(values)
                    min_val = min(values)
                    max_val = max(values)
                    variance = sum((x - avg) ** 2 for x in values) / len(values)
                    std_dev = variance ** 0.5

                    ws.cell(row=row, column=1, value=algorithm)
                    ws.cell(row=row, column=2, value=load)
                    ws.cell(row=row, column=3, value=avg)
                    ws.cell(row=row, column=4, value=min_val)
                    ws.cell(row=row, column=5, value=max_val)
                    ws.cell(row=row, column=6, value=std_dev)

                    row += 1

            # Add summary section
            row += 2
            summary_header = ws.cell(row=row, column=1, value="Summary by Algorithm")
            summary_header.font = header_font
            row += 1

            for algorithm in self.algorithms:
                all_values = []
                for load in self.loads:
                    all_values.extend(self.result.get((algorithm, load, metric_index), []))

                if not all_values:
                    continue  # Skip if no data

                avg = mean(all_values)
                min_val = min(all_values)
                max_val = max(all_values)
                variance = sum((x - avg) ** 2 for x in all_values) / len(all_values)
                std_dev = variance ** 0.5

                ws.cell(row=row, column=1, value=algorithm)
                ws.cell(row=row, column=3, value=avg)
                ws.cell(row=row, column=4, value=min_val)
                ws.cell(row=row, column=5, value=max_val)
                ws.cell(row=row, column=6, value=std_dev)

                row += 1

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

        self.wb.save(self.output)

        # --- Added Code for Saving Data to CSVs ---
        self.store_csvs(metrics)

        # --- Existing Code for Comparing QUEST ---
        self.compare_quest(metrics)

    def store_csvs(self, metrics: dict):
        """
        Stores the average, min, and max for each metric into separate CSV files.
        Each CSV file corresponds to a single metric and contains the Algorithm, Load, Average, Min, and Max.
        """
        # Define directory to store CSVs
        csv_dir = os.path.join(os.path.dirname(self.output), "csv_results")
        os.makedirs(csv_dir, exist_ok=True)

        for metric_name, metric_index in metrics.items():
            csv_filename = f"{metric_name.replace(' ', '_')}.csv"
            csv_path = os.path.join(csv_dir, csv_filename)

            with open(csv_path, mode='w', newline='') as csv_file:
                writer = csv.writer(csv_file)
                # Write headers
                writer.writerow(["Algorithm", "Load", "Average", "Min", "Max"])

                for algorithm in self.algorithms:
                    for load in self.loads:
                        values = self.result.get((algorithm, load, metric_index), [])
                        if not values:
                            continue  # Skip if no data

                        avg = mean(values)
                        min_val = min(values)
                        max_val = max(values)

                        writer.writerow([algorithm, load, avg, min_val, max_val])

        print(f"CSV files have been saved in the directory: {csv_dir}")

    def compare_quest(self, metrics: dict):
        """
        Compares the QUEST algorithm with other algorithms and prints the differences.
        """
        print("\n--- QUEST vs Other Algorithms Comparison ---\n")
        for metric_name, metric_index in metrics.items():
            print(f"Metric: {metric_name}")
            try:
                # Aggregate QUEST metrics across all loads
                quest_values = []
                for load in self.loads:
                    quest_values.extend(self.result[("QUEST", load, metric_index)])
                quest_avg = mean(quest_values)
                quest_min = min(quest_values)
                quest_max = max(quest_values)

                print(f"  QUEST - Average: {quest_avg:.4f}, Min: {quest_min:.4f}, Max: {quest_max:.4f}")

                for algorithm in self.algorithms:
                    if algorithm == "QUEST":
                        continue  # Skip QUEST itself

                    other_values = []
                    for load in self.loads:
                        other_values.extend(self.result.get((algorithm, load, metric_index), []))
                    if not other_values:
                        print(f"  {algorithm} has no data for this metric.")
                        continue
                    other_avg = mean(other_values)
                    other_min = min(other_values)
                    other_max = max(other_values)

                    lower_better_metrics = {"Data Age", "Energy", "Makespan", "Load"}
                    higher_better_metrics = {"Success Rate"}

                    if metric_name in lower_better_metrics:
                        avg_diff = other_avg - quest_avg
                        min_diff = other_min - quest_min
                        max_diff = other_max - quest_max
                        better_avg = "QUEST is better" if avg_diff > 0 else "Other is better"
                        better_min = "QUEST is better" if min_diff > 0 else "Other is better"
                        better_max = "QUEST is better" if max_diff > 0 else "Other is better"
                    elif metric_name in higher_better_metrics:
                        avg_diff = quest_avg - other_avg
                        min_diff = quest_min - other_min
                        max_diff = quest_max - other_max
                        better_avg = "QUEST is better" if avg_diff > 0 else "Other is better"
                        better_min = "QUEST is better" if min_diff > 0 else "Other is better"
                        better_max = "QUEST is better" if max_diff > 0 else "Other is better"
                    else:
                        avg_diff = other_avg - quest_avg
                        min_diff = other_min - quest_min
                        max_diff = other_max - quest_max
                        better_avg = "QUEST is better" if avg_diff > 0 else "Other is better"
                        better_min = "QUEST is better" if min_diff > 0 else "Other is better"
                        better_max = "QUEST is better" if max_diff > 0 else "Other is better"

                    print(f"  {algorithm} - Average Difference: {avg_diff:.4f} ({better_avg})")
                    print(f"             Min Difference: {min_diff:.4f} ({better_min})")
                    print(f"             Max Difference: {max_diff:.4f} ({better_max})")
            except KeyError:
                print("  QUEST data not available for this metric.")
            print()

    def create_sheet(self, sheet_name: str):
        if sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
        else:
            ws = self.wb.create_sheet(sheet_name)
        return ws

    def progress(self, percent: float, algorithm: str):
        arrow = '=' * int(round(percent * 100) - 1)
        spaces = ' ' * (100 - len(arrow))
        sys.stdout.write(f'\rProgress: [{arrow + spaces}] {int(percent * 100)}% [{algorithm}]')
        sys.stdout.flush()
