import random

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Font

from algorithms.Experiment import Experiment
from algorithms.NSGA3_AoI import NSGA3_AoI
from algorithms.QUEST import QUEST
from network.Network import Network

loads = [150]
dag_files = {
    150: [f"./workflow/Outputs/montage-150/dag-{i + 1}.json" for i in range(100)]
}
output = f"./motivation-1.xlsx"

samples = 100

if __name__ == '__main__':
    iterations = 100
    results: list[dict[str, float]] = []
    for i in range(iterations):
        results.append({
            "NSGA3" : 0.0,
            "QUEST": 0.0
        })

    for sample in range(samples):
        load = random.choice(loads)

        file = random.choice(dag_files[load])
        dag = Experiment.read_dag(file)
        network = Network.generate()

        Experiment.progress(sample / samples, "NSGA3")
        nsga3 = NSGA3_AoI(network, dag)
        nsga3.max_iterations = iterations
        nsga3.run()

        Experiment.progress(sample / samples, "QUEST")
        quest = QUEST(network, dag)
        quest.max_iterations = iterations
        quest.run()

        for i in range(iterations):
            if i < len(nsga3.per_iteration_objectives):
                results[i]["NSGA3"] += nsga3.per_iteration_objectives[i]
            else:
                results[i]["NSGA3"] += nsga3.per_iteration_objectives[len(nsga3.per_iteration_objectives) - 1]

        for i in range(iterations):
            if i < len(quest.per_iteration_objectives):
                results[i]["QUEST"] += quest.per_iteration_objectives[i]
            else:
                results[i]["QUEST"] += quest.per_iteration_objectives[len(quest.per_iteration_objectives) - 1]

        Experiment.progress((sample + 1) / samples, "QUEST")

    wb = load_workbook(output)

    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_font = Font(bold=True)

    ws = Experiment.create_sheet(wb, "Objective")

    row = 1
    for header in ["QUEST", "NSGA3"]:
        cell = ws.cell(row, 1, header)
        cell.fill = header_fill
        cell.font = header_font

        for i in range(iterations):
            ws.cell(row, i + 2, results[i][header] / samples)

        row += 1

    wb.save(output)
