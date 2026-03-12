import pandas as pd
import numpy as np
import os
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl import load_workbook

# Configuration
input_folder = r"C:\Users\mehrab\Desktop\QUEST\QUEST\result\v3\results-excel"
output_file = r"C:\Users\mehrab\Desktop\QUEST\QUEST\result\v4\charts_data.xlsx"

# Categories for the three types of charts
CATEGORIES = ["AoI", "Energy", "Makespan"]

# Scale mapping
SCALES = {
    '5': 'Small',
    '15': 'Medium',
    '25': 'Large'
}

# Regular expression to extract base name and number from filenames
FILENAME_PATTERN = re.compile(r"result-(.+)-(\d+)\.xlsx", re.IGNORECASE)

# Workload configurations with their loads
WORKLOAD_CONFIGS = {
    'alibaba': [1000, 2000],
    'cybershake': [30, 50, 100],
    'montage': [25, 50, 100],
    'epigenomics': [24, 46, 100],
    'inspiral': [30, 50, 100]
}

# Algorithm names (in order)
ALGORITHMS = ['QUEST', 'QUEST_ND', 'Fuzzy', 'NSGA3', 'MQGA', 'Greedy', 'MOPSO', 'ID3CO']

# Data locations in the Excel sheet for NORMALIZED values
NORMALIZED_DATA_LOCATIONS = [
    {'rows': (18, 25), 'cols': 'R'},  # AoI
    {'rows': (52, 59), 'cols': 'R'},  # Energy
    {'rows': (83, 90), 'cols': 'R'}  # Makespan
]

# End columns based on the number of loads
END_COLUMNS = {
    2: 'T',
    3: 'U',
    4: 'V',
}

# Workload order for output
WORKLOAD_ORDER = ['alibaba', 'inspiral', 'epigenomics', 'cybershake', 'montage']
DISPLAY_NAMES = {
    'alibaba': 'Alibaba',
    'montage': 'Montage',
    'cybershake': 'CyberShake',
    'epigenomics': 'Epigenomics',
    'inspiral': 'Inspiral'
}


def convert_col_letter_to_index(col_letter):
    """Convert column letter to numeric index (A=0, B=1, etc.)"""
    return ord(col_letter.upper()) - ord('A')


def extract_filename_info(filename):
    """Extract base name and number from filename."""
    match = FILENAME_PATTERN.match(filename)
    if match:
        return match.group(1), match.group(2)
    return None, None


def load_single_file_data(file_path, category_idx):
    """Load normalized data from a single Excel file for a specific category."""
    filename = os.path.basename(file_path)
    base_name, scale_num = extract_filename_info(filename)

    if not base_name:
        return None, None, None

    workload_key = base_name.lower()
    if workload_key not in WORKLOAD_CONFIGS:
        print(f"Unknown workload: {base_name}")
        return None, None, None

    loads = WORKLOAD_CONFIGS[workload_key]
    num_loads = len(loads)

    location = NORMALIZED_DATA_LOCATIONS[category_idx]
    start_row, end_row = location['rows']
    start_col_letter = location['cols']
    end_col_letter = END_COLUMNS.get(num_loads, 'T')

    start_col = convert_col_letter_to_index(start_col_letter)
    end_col = convert_col_letter_to_index(end_col_letter)

    try:
        raw_df = pd.read_excel(file_path, sheet_name='Charts', header=None)
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return None, None, None

    data_subset = raw_df.iloc[(start_row - 1):(end_row), start_col:(end_col + 1)]

    result = {}
    for idx in range(len(data_subset)):
        method_name = data_subset.iloc[idx, 0]
        if not isinstance(method_name, str):
            continue
        method_name = method_name.strip()

        if method_name == "QUEST_NDVFS":
            method_name = "QUEST_ND"
        elif method_name == "RL":
            method_name = "ID3CO"

        values = []
        for j in range(1, num_loads + 1):
            try:
                val = data_subset.iloc[idx, j]
                if pd.notna(val):
                    values.append(float(val))
                else:
                    values.append(np.nan)
            except (IndexError, ValueError):
                values.append(np.nan)

        result[method_name] = values

    return base_name, loads, result


def collect_all_data_for_scale(input_folder, scale_num, category_idx):
    """Collect data from all workloads for a specific scale and category."""
    input_path = Path(input_folder)
    pattern = f"result-*-{scale_num}.xlsx"
    files = list(input_path.glob(pattern))

    all_data = {}
    workload_loads = {}

    for file_path in files:
        base_name, loads, data = load_single_file_data(file_path, category_idx)
        if base_name and data:
            all_data[base_name.lower()] = data
            workload_loads[base_name.lower()] = loads

    return all_data, workload_loads


def write_sheet_with_two_header_rows(ws, all_data, workload_loads):
    existing_workloads = [w for w in WORKLOAD_ORDER if w in all_data]

    row1 = ['']
    for workload in existing_workloads:
        loads = workload_loads[workload]
        row1.append(DISPLAY_NAMES[workload])
        for _ in range(len(loads) - 1):
            row1.append('')

    row2 = ['']
    for workload in existing_workloads:
        loads = workload_loads[workload]
        for load in loads:
            row2.append(load)

    for col_idx, value in enumerate(row1, start=1):
        ws.cell(row=1, column=col_idx).value = value

    for col_idx, value in enumerate(row2, start=1):
        ws.cell(row=2, column=col_idx).value = value

    current_row = 3
    for algo in ALGORITHMS:
        row_data = [algo]
        for workload in existing_workloads:
            loads = workload_loads[workload]
            if algo in all_data[workload]:
                row_data.extend(all_data[workload][algo])
            else:
                row_data.extend([None] * len(loads))

        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=current_row, column=col_idx).value = value
        current_row += 1


def main():
    """Main function to update Excel data while preserving existing charts."""
    print("Starting Excel data update...")

    output_path = Path(output_file)

    if output_path.exists():
        print(f"Loading existing file: {output_file}")
        wb = load_workbook(output_file)
    else:
        print("Creating new workbook...")
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

    for scale_num, scale_name in SCALES.items():
        print(f"\nProcessing {scale_name} scale...")

        for category_idx, category in enumerate(CATEGORIES):
            print(f"  Processing {category}...")

            all_data, workload_loads = collect_all_data_for_scale(
                input_folder, scale_num, category_idx
            )

            if not all_data:
                print(f"    No data found for {category} - {scale_name}")
                continue

            sheet_name = f"{category}_{scale_name}"

            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=50):
                    for cell in row:
                        cell.value = None
            else:
                ws = wb.create_sheet(title=sheet_name)

            write_sheet_with_two_header_rows(ws, all_data, workload_loads)
            print(f"    Updated sheet: {sheet_name}")

    # Save workbook
    try:
        wb.save(output_file)
        print(f"\n✓ Successfully updated: {output_file}")
    except PermissionError:
        print("\n✗ Error: The Excel file is open! Please close it and run the script again.")


if __name__ == "__main__":
    main()