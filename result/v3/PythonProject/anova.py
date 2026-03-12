import pandas as pd
from openpyxl import load_workbook

# Configuration
file_path = '../results-excel/result-epigenomics-15.xlsx'
sheet_name = 'Data Age'
first_row = 52
end_row = 59

# Read the Excel file
df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

# Extract the data section (first_row to end_row, adjusting for 0-based indexing)
data_start_idx = first_row - 1  # Convert to 0-based index
data_end_idx = end_row - 1  # Convert to 0-based index

# Get the header row (row with "Algorithms", "1", "2", etc.)
header_row_idx = data_start_idx
algorithms_col_idx = 0  # Column A (0-based)

# Extract headers and data
headers = df.iloc[header_row_idx].tolist()
algorithms = df.iloc[data_start_idx + 1:data_end_idx + 2, algorithms_col_idx].tolist()

# Find all numeric load columns dynamically (columns with headers like "1", "2", "3", etc.)
load_columns = {}  # Dictionary: load_number -> list of column indices
for col_idx, header in enumerate(headers):
    if col_idx == 0:  # Skip the "Algorithms" column
        continue
    if header and isinstance(header, (int, float)):
        load_num = int(float(header))
        if load_num not in load_columns:
            load_columns[load_num] = []
        load_columns[load_num].append(col_idx)

# Sort load numbers
sorted_load_nums = sorted(load_columns.keys())

# Reorganize data as shown in picture 2
reorganized_data = []

# Add the header row
reorganized_data.append(['All data for chart'])

# Add column headers: "Algorithms", "Load 1", "Load 2", "Load 3", etc.
header_row = ['Algorithms'] + [f'Load {num}' for num in sorted_load_nums]
reorganized_data.append(header_row)

# Process each algorithm
for algo in algorithms:
    if pd.notna(algo) and algo != 'Algorithms':
        algo_idx = algorithms.index(algo)
        row_idx = data_start_idx + 1 + algo_idx

        # Collect values for each load number
        load_values = {}  # Dictionary: load_number -> list of values
        for load_num in sorted_load_nums:
            load_values[load_num] = []
            for col_idx in load_columns[load_num]:
                val = df.iloc[row_idx, col_idx]
                if pd.notna(val):
                    load_values[load_num].append(val)

        # Find the maximum number of values across all loads for this algorithm
        max_values = max(len(values) for values in load_values.values()) if load_values else 0

        # Add rows for each value
        for i in range(max_values):
            row = [algo]  # Algorithm name

            # Add values for each load
            for load_num in sorted_load_nums:
                if i < len(load_values[load_num]):
                    row.append(load_values[load_num][i])
                else:
                    row.append('')

            reorganized_data.append(row)

# Write the reorganized data to the Excel file
# Position: 3 rows after the main data ends (end_row + 3)
write_start_row = end_row + 6

# Load the workbook
wb = load_workbook(file_path)
ws = wb[sheet_name]

# Write the reorganized data
for i, row_data in enumerate(reorganized_data):
    for j, value in enumerate(row_data):
        # Convert to 1-based Excel row/column indexing
        ws.cell(row=write_start_row + i, column=j + 1, value=value)

# Save the workbook
wb.save(file_path)
print(f"Data reorganized and written to {file_path}")
print(f"New data starts at row {write_start_row}")
print(f"Total rows written: {len(reorganized_data)}")